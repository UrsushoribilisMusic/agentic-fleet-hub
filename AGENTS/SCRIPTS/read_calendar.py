import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\migue\Videos\MusicVideoOutput\video_tracker.xlsx', data_only=True)

ws = wb['Calendar']
headers = [c.value for c in ws[1]]
print('Columns:', headers)
print()
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    if vals[0] is None:
        continue
    print(' | '.join(str(v) if v is not None else '' for v in vals))
