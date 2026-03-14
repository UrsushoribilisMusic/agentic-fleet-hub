import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\migue\Videos\MusicVideoOutput\video_tracker.xlsx', data_only=True)

ws = wb['Shorts']
print('=== SHORTS ===')
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    sid = vals[0]
    if sid is None: continue
    title = (vals[1] or '')[:35]
    yt_v = vals[10] or 0; tt_v = vals[14] or 0; ig_v = vals[18] or 0
    print(f'S{sid:>2}: {title:35s} YT:{yt_v:>6} TT:{tt_v:>6} IG:{ig_v:>5}')

ws2 = wb['Videos']
headers2 = [c.value for c in ws2[1]]
print()
print('=== VIDEOS ===')
print('Columns:', headers2)
for row in ws2.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    vid = vals[0]
    if vid is None: continue
    title = (vals[1] or '')[:35]
    # Find view columns by index safely
    ncols = len(vals)
    yt_v = vals[10] if ncols > 10 else 0
    tt_v = vals[14] if ncols > 14 else 0
    ig_v = vals[18] if ncols > 18 else 0
    yt_v = yt_v or 0; tt_v = tt_v or 0; ig_v = ig_v or 0
    print(f'V{vid:>2}: {title:35s} YT:{yt_v:>6} TT:{tt_v:>6} IG:{ig_v:>5}')
