import openpyxl
import os

wb = openpyxl.load_workbook(r'C:\Users\migue\Videos\MusicVideoOutput\video_tracker.xlsx')
ws = wb['Chapters']

# Current columns (1-indexed):
# A=Ch, B=Act, C=Title, D=Characters, E=Summary, F=Substack URL,
# G=Lyrics File, H=Audio File, I=Video Files,
# J=TikTok Date, K=TikTok Views, L=TikTok Likes,
# M=Insta Date, N=Insta Views, O=Insta Likes,
# P=YT Short Date, Q=YT Short Views, R=YT Short Likes, S=Notes

# New column layout:
# A=Ch, B=Act, C=Title, D=Characters, E=Summary, F=Substack URL,
# G=Text, H=Prompts, I=Lyrics, J=Audio, K=Video Clips, L=Final Video,
# M=TikTok Date, N=TikTok Views, O=TikTok Likes,
# P=Insta Date, Q=Insta Views, R=Insta Likes,
# S=YT Short Date, T=YT Short Views, U=YT Short Likes, V=Notes

# First, collect existing data
existing = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    ch = row[0].value
    if ch is None:
        continue
    existing[ch] = {
        'act': row[1].value,
        'title': row[2].value,
        'characters': row[3].value,
        'summary': row[4].value,
        'url': row[5].value,
        'tt_date': row[9].value if len(row) > 9 else None,
        'tt_views': row[10].value if len(row) > 10 else None,
        'tt_likes': row[11].value if len(row) > 11 else None,
        'ig_date': row[12].value if len(row) > 12 else None,
        'ig_views': row[13].value if len(row) > 13 else None,
        'ig_likes': row[14].value if len(row) > 14 else None,
        'yt_date': row[15].value if len(row) > 15 else None,
        'yt_views': row[16].value if len(row) > 16 else None,
        'yt_likes': row[17].value if len(row) > 17 else None,
        'notes': row[18].value if len(row) > 18 else None,
    }

# New chapters data
new_chapters = {
    20: {
        'title': 'The Ethical Barrier',
        'characters': 'P2, P3',
        'summary': 'Paul pitches the brain-reading job to Erica. She demands ethics paperwork and refuses to proceed without proper documentation.',
        'url': 'https://scifibymiguel.substack.com/p/chapter-20-the-ethical-barrier',
    },
    21: {
        'title': 'The Paper Architect',
        'characters': 'P1',
        'summary': 'Zheng instructs Lena to obtain forged documents from Panama. Lena conflicted about manipulating Erica.',
        'url': 'https://scifibymiguel.substack.com/p/the-paper-architect',
    },
    22: {
        'title': 'Building Trust',
        'characters': 'P1, P3',
        'summary': 'Lena presents fake "Panama Dossiers" to Erica at a bar. Genuine attraction builds while Lena deceives her.',
        'url': 'https://scifibymiguel.substack.com/p/building-trust',
    },
    23: {
        'title': 'The Strategy of the Hook',
        'characters': 'Inspector Pinto, Captain Meier',
        'summary': 'Inspector Pinto arrives in Bern. Police analyze the trio. Plan: let them fly to SF, then pressure Lena to flip as controlled asset.',
        'url': 'https://scifibymiguel.substack.com/p/chapter-23-the-strategy-of-the-hook',
    },
    24: {
        'title': 'The Flight',
        'characters': 'P1, P2, P3',
        'summary': 'The trio flies Zurich to SF in business class. Erica confronts Lena about her true allegiances. Surveillance operative on the plane.',
        'url': 'https://scifibymiguel.substack.com/p/chapter-25-the-flight',
    },
}

# Add new chapters to existing
for ch_num, data in new_chapters.items():
    if ch_num not in existing:
        existing[ch_num] = {
            'act': None, 'characters': data['characters'],
            'summary': data['summary'], 'url': data['url'],
            'title': data['title'],
            'tt_date': None, 'tt_views': None, 'tt_likes': None,
            'ig_date': None, 'ig_views': None, 'ig_likes': None,
            'yt_date': None, 'yt_views': None, 'yt_likes': None,
            'notes': None,
        }
    else:
        # Update title, summary, URL for existing entries if empty
        if not existing[ch_num].get('title'):
            existing[ch_num]['title'] = data['title']
        if not existing[ch_num].get('summary'):
            existing[ch_num]['summary'] = data['summary']
        if not existing[ch_num].get('url'):
            existing[ch_num]['url'] = data['url']
        if not existing[ch_num].get('characters'):
            existing[ch_num]['characters'] = data['characters']

# Scan directories for file status
base = r'C:\Users\migue\Videos\StoryVideos'

def scan_chapter(ch_num):
    ch_dir = os.path.join(base, f'Chapter{ch_num}')
    if not os.path.isdir(ch_dir):
        return {'text': '', 'prompts': '', 'lyrics': '', 'audio': '', 'clips': 0, 'final': ''}

    files = os.listdir(ch_dir)
    text = [f for f in files if f.lower().endswith('.txt') and 'text' in f.lower()]
    prompts = [f for f in files if 'prompt' in f.lower()]
    lyrics = [f for f in files if 'lyrics' in f.lower() or 'lyric' in f.lower()]
    audio = [f for f in files if f.lower().endswith(('.mp3', '.wav', '.m4a', '.flac'))]
    clips = [f for f in files if f.lower().endswith('.mp4') and 'final' not in f.lower() and 'intro' not in f.lower()]
    final = [f for f in files if f.lower().endswith('.mp4') and 'final' in f.lower()]

    return {
        'text': 'Yes' if text else '',
        'prompts': 'Yes' if prompts else '',
        'lyrics': 'Yes' if lyrics else '',
        'audio': 'Yes' if audio else '',
        'clips': len(clips) if clips else 0,
        'final': 'Yes' if final else '',
    }

# Clear existing data rows
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Write new headers
headers = ['Ch', 'Act', 'Title', 'Characters', 'Summary', 'Substack URL',
           'Text', 'Prompts', 'Lyrics', 'Audio', 'Video Clips', 'Final Video',
           'TikTok Date', 'TikTok Views', 'TikTok Likes',
           'Insta Date', 'Insta Views', 'Insta Likes',
           'YT Short Date', 'YT Short Views', 'YT Short Likes', 'Notes']

from openpyxl.styles import Font, Alignment, PatternFill

header_font = Font(bold=True)
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font

# Write data rows
row_num = 2
for ch_num in sorted(existing.keys()):
    data = existing[ch_num]
    scan = scan_chapter(ch_num)

    ws.cell(row=row_num, column=1, value=ch_num)
    ws.cell(row=row_num, column=2, value=data.get('act'))
    ws.cell(row=row_num, column=3, value=data.get('title'))
    ws.cell(row=row_num, column=4, value=data.get('characters'))
    ws.cell(row=row_num, column=5, value=data.get('summary'))
    ws.cell(row=row_num, column=6, value=data.get('url'))
    ws.cell(row=row_num, column=7, value=scan['text'])
    ws.cell(row=row_num, column=8, value=scan['prompts'])
    ws.cell(row=row_num, column=9, value=scan['lyrics'])
    ws.cell(row=row_num, column=10, value=scan['audio'])
    ws.cell(row=row_num, column=11, value=scan['clips'] if scan['clips'] > 0 else '')
    ws.cell(row=row_num, column=12, value=scan['final'])
    ws.cell(row=row_num, column=13, value=data.get('tt_date'))
    ws.cell(row=row_num, column=14, value=data.get('tt_views'))
    ws.cell(row=row_num, column=15, value=data.get('tt_likes'))
    ws.cell(row=row_num, column=16, value=data.get('ig_date'))
    ws.cell(row=row_num, column=17, value=data.get('ig_views'))
    ws.cell(row=row_num, column=18, value=data.get('ig_likes'))
    ws.cell(row=row_num, column=19, value=data.get('yt_date'))
    ws.cell(row=row_num, column=20, value=data.get('yt_views'))
    ws.cell(row=row_num, column=21, value=data.get('yt_likes'))
    ws.cell(row=row_num, column=22, value=data.get('notes'))

    row_num += 1

# Set column widths
widths = {'A': 4, 'B': 5, 'C': 25, 'D': 15, 'E': 60, 'F': 45,
          'G': 6, 'H': 8, 'I': 7, 'J': 7, 'K': 11, 'L': 11,
          'M': 12, 'N': 11, 'O': 11, 'P': 12, 'Q': 11, 'R': 11,
          'S': 12, 'T': 11, 'U': 11, 'V': 20}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

wb.save(r'C:\Users\migue\Videos\MusicVideoOutput\video_tracker.xlsx')
print('Chapters sheet updated successfully!')
print(f'Total chapters: {len(existing)} (Ch0-Ch{max(existing.keys())})')
