import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\migue\Videos\MusicVideoOutput\video_tracker.xlsx', data_only=True)
ws = wb['Chapters']

headers = [c.value for c in ws[1]]
print('Columns:', headers)
print()
print(f'{"Ch":>3} {"Title":25s} {"Text":5s} {"Prm":4s} {"Lyr":4s} {"Aud":4s} {"Clips":5s} {"Final":5s}')
print('-' * 60)
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    if vals[0] is None: continue
    ch = vals[0]
    title = (vals[2] or '')[:24]
    text = vals[6] or ''
    prompts = vals[7] or ''
    lyrics = vals[8] or ''
    audio = vals[9] or ''
    clips = vals[10] or ''
    final = vals[11] or ''
    print(f'{ch:>3} {title:25s} {text:5s} {prompts:4s} {lyrics:4s} {audio:4s} {str(clips):5s} {final:5s}')
